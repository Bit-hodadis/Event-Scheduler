# management/commands/import_2016_tax_plan.py
import decimal

from django.core.management.base import BaseCommand
from django.db import transaction
from openpyxl import load_workbook

from address.models import Address
from plan.models import AnnualTaxPlan, HigherAuthorityMonthlyTaxPlan
from tax.models import TaxType
from users.models import CustomUser

MONTH_MAPPING = {
    "November": 11,
    "December": 12,
    "Junary": 1,  # Assuming this is January (typo in Excel)
    "Febrary": 2,  # Assuming this is February (typo in Excel)
    "March": 3,
    "April": 4,
    "May": 5,
    "July": 7,
    "Oguste": 8,  # Assuming this is August
    "Jun": 6,
    "octobre": 10,  # Assuming this is October
    # "In general": None,  # Annual total, not a month
}


class Command(BaseCommand):
    help = "Import 2016 tax plan data from Excel file"

    def add_arguments(self, parser):
        parser.add_argument("file_path", type=str, help="Path to the Excel file")

    def handle(self, *args, **options):
        file_path = options["file_path"]

        try:
            wb = load_workbook(filename=file_path, data_only=True)
            sheet = wb["Karoora GI & MQ 2016 Ji'aan"]  # First sheet

            # Get or create tax types
            ordinary_tax_type = TaxType.objects.all().first()

            # with transaction.atomic():
            for row in sheet.iter_rows(min_row=6, values_only=True):
                # Skip empty rows or summary rows
                if not row[1] or row[1] in ("sum", "Ida'ama"):
                    continue

                region_name = row[1].strip()
                # annual_total = decimal.Decimal(str(row[-1])) if row[-1] else None
                users_data = CustomUser.objects.filter(is_superuser=True).first()

                oromia = Address.objects.filter(parent=None).first()

                # Get or create address
                address, _ = Address.objects.get_or_create(
                    name=region_name,
                    parent=oromia,
                    created_by=users_data,  # Assuming superuser
                )

                # Create annual tax plan
                annual_plan, created = AnnualTaxPlan.objects.get_or_create(
                    address=address,
                    year=2016,
                    defaults={
                        # "annual_amount": annual_total,
                        "created_by": users_data,
                    },
                )
                print(row, HigherAuthorityMonthlyTaxPlan.objects.all())

                # Process monthly ordinary plan data (columns C-N)
                i = 1
                for month_name in row[2:14]:  # Month names from header row
                    month_number = i
                    i += 1
                    if not month_number:
                        continue

                    amount = month_name  # +2 to skip No and Name columns
                    if not amount:
                        continue

                    print(amount, "it is an amount")

                    HigherAuthorityMonthlyTaxPlan.objects.update_or_create(
                        annual_tax_plan=annual_plan,
                        tax_type=ordinary_tax_type,
                        month=month_number,
                        defaults={
                            "monthly_amount": amount,
                            "created_by": users_data,
                        },
                    )

                self.stdout.write(self.style.SUCCESS("Successfully imported data"))

        except Exception as e:
            self.stdout.write(self.style.ERROR(f"Error importing data: {str(e)}"))
            # raise
